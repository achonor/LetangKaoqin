import os
import shutil
import xlrd
import openpyxl.cell
from openpyxl import utils
from openpyxl.styles import numbers
from Scripts.Logic import Functions
from Scripts.Data.Configs import Configs
from datetime import datetime, timedelta
from Scripts.Data.CommonDatas import CommonDatas
from openpyxl.styles import Color, Font, Alignment, PatternFill, Border, Side, Protection, colors

const_fill_red = PatternFill('solid', fgColor='FF0000')
const_fill_yellow = PatternFill('solid', fgColor='FFC000')
const_fill_bule = PatternFill('solid', fgColor='0000FF')
const_fill_purple = PatternFill('solid', fgColor='7030A0')

# 边框样式
const_border = Border(left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"))
# 对齐
const_alignment = Alignment(horizontal='center', vertical='center')
# 字体
const_font = Font(name="Arial", sz=10, bold=False)
const_font_red = Font(name="Arial", sz=10, bold=False, color=colors.RED)
const_bold_font = Font(name="Arial", sz=10, bold=True)
const_bold_font_red = Font(name="Arial", sz=10, bold=True, color=colors.RED)
const_title_font = Font(name="宋体", sz=22, bold=True)



def copy_template(template_path, out_dir, out_name_template):
    start_date = CommonDatas.get_instance().start_date
    out_name = out_name_template.format(start_date.month, os.path.splitext(template_path)[1])
    out_path = os.path.join(out_dir, out_name)
    shutil.copyfile(template_path, out_path)
    return out_path

def export_detail_table(template_path, out_dir):
    start_date = CommonDatas.get_instance().start_date
    out_path = os.path.join(out_dir, "{0:02d}月考勤明细.xlsx".format(start_date.month))
    #out_path = copy_template(template_path, out_dir, "{0:02d}月考勤明细{1}")
    xlsx_table = openpyxl.Workbook()
    xlsx_sheet = xlsx_table.get_active_sheet()
    #成员信息
    member_dict = CommonDatas.get_instance().get_all_member_dict()
    member_list = list(member_dict.values())
    #初始表格样式
    sheet_row, sheet_column = init_detail_sheet(xlsx_sheet, len(member_list), CommonDatas.get_instance().days_number)
    #成员数据
    for index, member in enumerate(member_list):
        row = 5 + index
        member.calc_all_data()
        #序号
        xlsx_sheet.cell(row, 1).value = index + 1
        #名字
        xlsx_sheet.cell(row, 2).value = member.name
        #打卡情况
        for day in range(1, CommonDatas.get_instance().days_number + 1):
            colum1 = 2 + 2 * day - 1    #上午
            colum2 = colum1 + 1         #下午
            cur_date = CommonDatas.get_instance().create_datetime(day=day)
            xlsx_cell1 = xlsx_sheet.cell(row, colum1)
            xlsx_cell2 = xlsx_sheet.cell(row, colum2)
            #初始化
            xlsx_cell1.value = "√"
            xlsx_cell2.value = "√"

            if not member.in_company(day):
                #没入职
                xlsx_sheet.merge_cells(None, row, colum1, row, colum2)
                xlsx_cell1.value = "未入职"
            elif not Configs.get_instance().get_is_work_time(day):
                #休息
                xlsx_cell1.value = "休息"
                xlsx_cell2.value = "休息"
            elif 0 < member.is_late_or_early(day):
                #迟到或者早退
                duration, late, early = member.get_late_and_early_time(day)
                if 1440 == duration or 480 == duration:
                    xlsx_sheet.merge_cells(None, row, colum1, row, colum2)
                    xlsx_cell1.fill = const_fill_purple
                    xlsx_cell1.value = "缺卡"
                else:
                    if 0 < late:
                        xlsx_cell1.fill = const_fill_yellow
                        xlsx_cell1.value = "迟到{0}分钟".format(late)
                    if 0 < early:
                        xlsx_cell2.fill = const_fill_yellow
                        xlsx_cell2.value = " 早退{0}分钟".format(early)
            #计算请假
            leave_duration, am_duration, pm_duration = member.get_leave_duration(day)
            work_day_duration = CommonDatas.get_instance().get_work_day_duration()
            if leave_duration == work_day_duration:
                #全天请假
                xlsx_cell1.fill = const_fill_red
                xlsx_cell2.fill = const_fill_red
                xlsx_sheet.merge_cells(None, row, colum1, row, colum2)
                xlsx_cell1.value = "请假" + Functions.timedelta2string(leave_duration)
            else:
                if 0 < am_duration.total_seconds():
                    xlsx_cell1.fill = const_fill_red
                    xlsx_cell1.value = "请假" + Functions.timedelta2string(am_duration)
                if 0 < pm_duration.total_seconds():
                    xlsx_cell2.fill = const_fill_red
                    xlsx_cell2.value = "请假" + Functions.timedelta2string(pm_duration)

        cloum_offset = 2 + 2 * CommonDatas.get_instance().days_number + 1
        #请假次数
        xlsx_sheet.cell(row, cloum_offset).value = member.get_leave_count()
        #请假小时
        xlsx_sheet.cell(row, cloum_offset + 1).value = Functions.timedelta2string(member.get_leave_duration())
        #迟到次数
        xlsx_sheet.cell(row, cloum_offset + 2).value = member.late_or_early_count
        #缺卡次数
        xlsx_sheet.cell(row, cloum_offset + 3).value = member.lack_checkin_count
        #罚款金额
        xlsx_sheet.cell(row, cloum_offset + 4).value = member.late_or_early_fine_number
    xlsx_table.save(out_path)

def export_summary_table(template_path, out_dir):
    #out_path = copy_template(template_path, out_dir, "{0:02d}月考勤汇总{1}")
    start_date = CommonDatas.get_instance().start_date
    end_date = CommonDatas.get_instance().end_date
    out_path = os.path.join(out_dir, "{0:02d}月考勤汇总.xlsx".format(start_date.month))
    xlsx_table = openpyxl.Workbook()
    xlsx_sheet_summary = xlsx_table.get_active_sheet()
    xlsx_sheet_summary.title = "考勤汇总"
    xlsx_sheet_other = xlsx_table.create_sheet("请假加迟到次数明细")
    xlsx_sheet_member = xlsx_table.create_sheet("邮箱和卡号")
    #成员信息
    member_dict = CommonDatas.get_instance().get_all_member_dict()
    member_list = list(member_dict.values())
    member_number = len(member_list)
    #初始化考勤汇总sheet
    init_summary_sheet(xlsx_sheet_summary, member_number)
    #初始化请假加迟到次数明细
    init_other_sheet(xlsx_sheet_other, member_number)
    #初始化邮箱卡号
    init_member_sheet(xlsx_sheet_member, member_number)
    #生成考勤汇总sheet
    for index, member in enumerate(member_list):
        row = index + 2
        #序号
        xlsx_sheet_summary.cell(row, 1).value = index + 1
        #入职时间
        xlsx_sheet_summary.cell(row, 2).value = member.in_date.strftime("%Y.%m.%d")
        #姓名
        xlsx_sheet_summary.cell(row, 3).value = member.name
        #职位
        xlsx_sheet_summary.cell(row, 4).value = member.profession
        #部门
        xlsx_sheet_summary.cell(row, 5).value = member.department_1
        #事业部
        xlsx_sheet_summary.cell(row, 6).value = member.department_2
        #考勤
        attendance = member.get_attendance()
        if 0 != attendance:
            xlsx_sheet_summary.cell(row, 7).value = attendance
        if member.in_date.year == start_date.year and member.in_date.month + 1 == start_date.month and 15 < member.in_date.day:
            xlsx_sheet_summary.cell(row, 7).value = "需手动计算上月考勤 + " + str(attendance)
            #离职时间
        if start_date <= member.out_date and member.out_date <= end_date:
            xlsx_sheet_summary.cell(row, 8).value = member.out_date.strftime("%Y.%m.%d")

        #请假和迟到次数总和
        leave_late_early_count = member.get_leave_count() + member.late_or_early_count
        if 0 != leave_late_early_count:
            xlsx_sheet_summary.cell(row, 9).value = leave_late_early_count
    # 生成考勤汇总sheet结束

    #生成请假加迟到次数明细sheet
    for index, member in enumerate(member_list):
        row = index + 2
        #姓名
        xlsx_sheet_other.cell(row, 1).value = member.name
        #事假
        if 0 < member.leave_durations.total_seconds():
            xlsx_sheet_other.cell(row, 2).value = Functions.timedelta2string(member.leave_durations)
        #病假
        #年假
        #迟到
        if 0 < member.late_or_early_count:
            xlsx_sheet_other.cell(row, 5).value = member.late_or_early_count
        #请假时间
        xlsx_sheet_other.cell(row, 6).value = Functions.leave_string2ecxel_text(member.get_leave_text_list())

    #生成邮箱卡号sheet
    for index, member in enumerate(member_list):
        row = index + 2
        #姓名
        xlsx_sheet_member.cell(row, 1).value = member.name
        #卡号
        xlsx_sheet_member.cell(row, 2).value = member.bank_card
        #邮箱
        xlsx_sheet_member.cell(row, 3).value = member.email
    xlsx_table.save(out_path)


def init_detail_sheet(sheet, member_number, day_number):
    #行列数
    sheet_row = 4 + member_number
    sheet_column = 2 + 2 * day_number + 5
    #年月信息
    start_date = CommonDatas.get_instance().start_date
    #行高列宽
    sheet.row_dimensions[1].height = 45
    for index in range(2, sheet_row + 1):
        sheet.row_dimensions[index].height = 20
    sheet.column_dimensions["A"].width = 5
    sheet.column_dimensions["B"].width = 8.5
    for index in range(3, sheet_column - 4):
        index_str = utils.get_column_letter(index)
        sheet.column_dimensions[index_str].width = 8.5
    sheet.column_dimensions[utils.get_column_letter(sheet_column-4)].width = 9
    sheet.column_dimensions[utils.get_column_letter(sheet_column-3)].width = 9
    sheet.column_dimensions[utils.get_column_letter(sheet_column-2)].width = 9
    sheet.column_dimensions[utils.get_column_letter(sheet_column-1)].width = 11
    sheet.column_dimensions[utils.get_column_letter(sheet_column)].width = 13
    #标题
    sheet.merge_cells(None, 1, 1, 1, sheet_column)
    for index in range(1, sheet_column + 1):
        sheet.cell(1, index).border = const_border
    set_cell_fromat(sheet.cell(1, 1), font=const_title_font, border=None)
    sheet.cell(1, 1).value = "{0}月考勤表（黄色填充代表迟到，蓝色填充代表未打卡，红色填充代表请假)".format(start_date.month)
    #表头
    sheet.merge_cells(None, 2, 1, 3, 2)
    #日历星期
    for day in range(1, 31 + 1):
        column = 2 + 2 * day - 1
        day_cell = sheet.cell(2, column)
        week_cell = sheet.cell(3, column)
        set_cell_fromat(day_cell)
        set_cell_fromat(week_cell)
        if day <= CommonDatas.get_instance().days_number:
            cur_date = CommonDatas.get_instance().create_datetime(day=day)
            sheet.merge_cells(None, 2, column, 2, column + 1)
            day_cell.number_format = 'm月d日'
            day_cell.value = cur_date
            #星期
            sheet.merge_cells(None, 3, column, 3, column + 1)
            week_cell.value = Functions.weekday2string(cur_date.weekday())
        else:
            pass
            #xlsx_sheet.delete_cols(column, 2)
    #统计头部
    sheet.merge_cells(None, 2, sheet_column - 4, 3, sheet_column)
    #序号，姓名
    set_cell_fromat(sheet.cell(4, 1))
    sheet.cell(4, 1).value = "序号"
    set_cell_fromat(sheet.cell(4, 2))
    sheet.cell(4, 2).value = "名字"
    #上下班打卡
    for index in range(3, sheet_column - 4):
        set_cell_fromat(sheet.cell(4, index))
        sheet.cell(4, index).value = Functions.three_param_operator(1 == index % 2, "上班打卡", "下班打卡")
    #请假次数
    set_cell_fromat(sheet.cell(4, sheet_column - 4))
    sheet.cell(4, sheet_column - 4).value = "请假次数"
    #请假小时
    set_cell_fromat(sheet.cell(4, sheet_column - 3))
    sheet.cell(4, sheet_column - 3).value = "请假小时"
    #迟到次数
    set_cell_fromat(sheet.cell(4, sheet_column - 2))
    sheet.cell(4, sheet_column - 2).value = "迟到次数"
    #忘打卡次数
    set_cell_fromat(sheet.cell(4, sheet_column - 1))
    sheet.cell(4, sheet_column - 1).value = "忘打卡次数"
    #罚款金额（元）
    set_cell_fromat(sheet.cell(4, sheet_column))
    sheet.cell(4, sheet_column).value = "罚款金额（元）"
    #成员格式
    for index in range(1, member_number + 1):
        row = 4 + index
        #序号
        set_cell_fromat(sheet.cell(row, 1))
        #名字
        set_cell_fromat(sheet.cell(row, 2))
        #打卡信息
        for idx in range(3, sheet_column - 4):
            set_cell_fromat(sheet.cell(row, idx))
        #统计信息
        set_cell_fromat(sheet.cell(row, sheet_column - 4))
        set_cell_fromat(sheet.cell(row, sheet_column - 3))
        set_cell_fromat(sheet.cell(row, sheet_column - 2))
        set_cell_fromat(sheet.cell(row, sheet_column - 1))
        set_cell_fromat(sheet.cell(row, sheet_column))
    return sheet_row, sheet_column


def init_summary_sheet(sheet, member_number):
    sheet_column = 10
    first_row_width = [6.5, 12, 8.5, 10, 13, 13, 11, 12, 20, 22]
    first_row_value = ["工号", "入职时间", "姓名", "职位", "部门", "事业部", "考勤", "离职时间（最后工作日）", "所有假和迟到次数总和", "符合扣除绩效情况"]
    #第一排
    sheet.row_dimensions[1].height = 38
    for index in range(1, sheet_column + 1):
        if 7 == index or 8 == index or 9 == index:
            set_cell_fromat(sheet.cell(1, index), font=const_bold_font_red)
        else:
            set_cell_fromat(sheet.cell(1, index), font=const_bold_font)
        sheet.cell(1, index).value = first_row_value[index - 1]
        sheet.column_dimensions[utils.get_column_letter(index)].width = first_row_width[index - 1]

    #成员
    for index in range(1, member_number + 1):
        row = index + 1
        sheet.row_dimensions[row].height = 24
        for idx in range(1, sheet_column + 1):
            if 7 == idx or 8 == idx or 9 == idx:
                set_cell_fromat(sheet.cell(row, idx), font=const_font_red)
            else:
                set_cell_fromat(sheet.cell(row, idx))

def init_other_sheet(sheet, member_number):
    sheet_column = 6
    first_row_width = [8.5, 11, 8.5, 8.5, 8.5, 80]
    first_row_value = ["姓名", "事假", "病假", "年假", "迟到", "事件"]
    #第一排
    sheet.row_dimensions[1].height = 20
    for index in range(1, sheet_column + 1):
        set_cell_fromat(sheet.cell(1, index), font=const_bold_font)
        sheet.cell(1, index).value = first_row_value[index - 1]
        sheet.column_dimensions[utils.get_column_letter(index)].width = first_row_width[index - 1]
    #成员
    for index in range(1, member_number + 1):
        row = index + 1
        sheet.row_dimensions[row].height = 20
        for idx in range(1, sheet_column + 1):
            set_cell_fromat(sheet.cell(row, idx))

def init_member_sheet(sheet, member_number):
    sheet_column = 3
    first_row_width = [11, 22, 28]
    first_row_value = ["姓名", "卡号", "邮箱"]
    # 第一排
    sheet.row_dimensions[1].height = 20
    for index in range(1, sheet_column + 1):
        set_cell_fromat(sheet.cell(1, index), font=const_bold_font)
        sheet.cell(1, index).value = first_row_value[index - 1]
        sheet.column_dimensions[utils.get_column_letter(index)].width = first_row_width[index - 1]
    # 成员
    for index in range(1, member_number + 1):
        row = index + 1
        sheet.row_dimensions[row].height = 20
        for idx in range(1, sheet_column + 1):
            set_cell_fromat(sheet.cell(row, idx))

def set_cell_fromat(cell, font=const_font, border=const_border, alignment=const_alignment):
    if font is not None:
        cell.font = font
    if border is not None:
        cell.border = border
    if alignment is not None:
        cell.alignment = alignment